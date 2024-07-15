from typing import Iterable

import itertools
import locale
from datetime import datetime

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import Statement, StatementLine, generate_transaction_id
from openpyxl import load_workbook
from typing import Any, List

import logging


def take(n: int, iterable: Iterable[Any]):
    return list(itertools.islice(iterable, n))


class OKQ8Plugin(Plugin):
    """OKQ8 Bank <https://www.okq8.se/bank/>"""

    def get_parser(self, filename: str) -> "OKQ8Parser":
        return OKQ8Parser(filename)


class OKQ8Parser(StatementParser[Any]):
    DATE_FORMAT = "%d %b %Y"
    TRANSACTION_DATE = 0
    TRANSACTION_DESCRIPTION = 2
    TRANSACTION_AMOUNT = 5
    TRANSACTION_CREDIT = 3
    TRANSACTION_DEBIT = 4

    def __init__(self, filename: str) -> None:
        super().__init__()
        self.filename = filename
        self.sheet = load_workbook(filename=filename, read_only=True).active
        locale.setlocale(locale.LC_TIME, "sv_SE.UTF-8")

    def parse(self) -> Statement:
        """Main entry point for parsers

        super() implementation will call to split_records and parse_record to
        process the file.
        """

        assert self.sheet is not None
        self.statement.account_id = "OKQ8"  # TODO: Config?
        self.statement.currency = "SEK"
        self.statement.bank_id = "OKQ8"

        logging.info(f"SELF f{self.statement}")
        rows = self.sheet.iter_rows()
        # Get headers
        (header,) = list(take(1, rows))
        logging.info(f"HEADERS {[c.value for c in header]}")
        assert [
            "Datum",
            "Typ",
            "Beskrivning",
            "In på konto",
            "Ut från konto",
            "Originalt belopp",
            "Valuta",
            "Kurs",
        ] == [c.value for c in header]

        self.rows = list(rows)

        return super().parse()

    def split_records(self) -> Iterable[Any]:
        """Return iterable object consisting of a line per transaction"""
        for row in self.rows:
            yield [c.value for c in row]

    def parse_record(self, line: List[Any]) -> StatementLine:
        """Parse given transaction line and return StatementLine object"""
        stmt_line = StatementLine()

        date = line[self.TRANSACTION_DATE]

        date = datetime.strptime(date, self.DATE_FORMAT)
        description = line[self.TRANSACTION_DESCRIPTION]
        amount = line[self.TRANSACTION_AMOUNT]
        debit = line[self.TRANSACTION_DEBIT]
        credit = line[self.TRANSACTION_CREDIT]

        trntype, amount = self.get_type_and_amount(debit, credit, amount)

        logging.info(f"TRNTYPE={trntype}, amount={amount}, desc={description}")

        stmt_line.date = date
        stmt_line.memo = description
        stmt_line.amount = amount
        stmt_line.trntype = trntype

        stmt_line.id = generate_transaction_id(stmt_line)

        return stmt_line

    @staticmethod
    def get_type_and_amount(debit, credit, amount):
        if credit is not None:
            return "CREDIT", credit
        elif debit is not None:
            return "DEBIT", debit
        else:
            return "OTHER", amount
